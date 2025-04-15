from openpyxl import load_workbook

# 打开 Excel 文件
workbook = load_workbook('SC_modified1.xlsx')

# 处理每个 sheet
for sheetname in workbook.sheetnames:
    sheet = workbook[sheetname]
    
    # 遍历每一行，从最后一行开始，避免删除行导致索引错误
    for row_index in range(sheet.max_row, 0, -1):
        # 检查当前行的每个单元格值
        for cell in sheet[row_index]:
            if cell.value == 'CAS':
                break
        else:
            # 如果当前行中没有单元格值为 "CAS"，则删除该行
            sheet.delete_rows(row_index)

# 保存修改后的 Excel 文件
workbook.save('SC_modified1_updated.xlsx')
