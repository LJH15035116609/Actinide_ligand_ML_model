import openpyxl

def extract_contents_from_E(sheet):
    extracted_contents = []
    for cell in sheet['E']:
        cell_value = cell.value
        if isinstance(cell_value, str) and 'M' in cell_value:
            content = cell_value.replace('M', '').strip()
            if content:  # 确保内容非空
                try:
                    content = float(content)  # 尝试将内容转换为数字格式
                    extracted_contents.append(content)
                except ValueError:
                    pass

    return extracted_contents

# 打开 Excel 文件
workbook = openpyxl.load_workbook('SC_modified3 - 副本.xlsx')

# 创建一个新的 Excel 文件来保存提取的数据
output_workbook = openpyxl.Workbook()

# 遍历每个 sheet
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    extracted_contents = extract_contents_from_E(sheet)

    # 创建一个新的工作表来保存提取的数据
    output_sheet = output_workbook.create_sheet(title=sheet_name)

    # 将提取的数据写入到新的工作表中
    for content in extracted_contents:
        output_sheet.append([content])

    # 设置单元格格式为数字
    for column_cells in output_sheet.columns:
        for cell in column_cells:
            cell.number_format = '0.00'

# 删除默认的工作表
output_workbook.remove(output_workbook.active)

# 保存新的 Excel 文件
output_workbook.save('M.xlsx')
