from openpyxl import load_workbook
from openpyxl import Workbook

# 打开原始文件
input_file = "SC_modified1_updated.xlsx"
wb = load_workbook(input_file)

# 创建一个新的Excel文件
output_wb = Workbook()

# 遍历每个工作表
for sheet_name in wb.sheetnames:
    # 选择当前工作表
    ws = wb[sheet_name]
    
    # 创建一个新的工作表，用于保存提取的内容
    output_ws = output_wb.create_sheet(title=sheet_name)
    
    # 记录提取内容的行数
    row_count = 1
    
    # 遍历每个单元格
    for row in ws.iter_rows():
        for cell in row:
            # 检查单元格的值是否符合条件
            if cell.value == "CAS":
                # 获取相邻单元格的内容（列数+1，行数不变）
                next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                # 将提取的内容写入新的工作表中的下一行
                output_ws.cell(row=row_count, column=1, value=next_cell.value)
                row_count += 1

# 删除默认创建的第一个工作表
output_wb.remove(output_wb["Sheet"])

# 保存新的Excel文件
output_file = "extracted_data.xlsx"
output_wb.save(output_file)


