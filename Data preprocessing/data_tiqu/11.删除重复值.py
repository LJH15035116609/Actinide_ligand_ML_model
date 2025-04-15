import openpyxl

def process_sheet(sheet):
    unique_rows = {}  # 用字典存储唯一的行，键为B、D、E列的值组成的元组，值为行号
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):  # 从第2行开始遍历
        key = (row[1], row[3], row[4])  # 获取B、D、E列的值组成的元组作为键
        if key not in unique_rows:
            unique_rows[key] = row_index
        elif unique_rows[key] > row_index:  # 只保留行数最小的行
            unique_rows[key] = row_index

    rows_to_keep = set(unique_rows.values())

    for row_index in range(sheet.max_row, 1, -1):  # 倒序删除以避免索引错误
        if row_index not in rows_to_keep:
            sheet.delete_rows(row_index)

def process_workbook(input_file, output_file):
    workbook = openpyxl.load_workbook(input_file)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        process_sheet(sheet)

    workbook.save(output_file)

# 调用函数处理 SC_modified4.xlsx 文件
process_workbook('SC_modified4.xlsx', 'SC_modified5.xlsx')
