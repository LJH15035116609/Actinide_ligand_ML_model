import os
from openpyxl import load_workbook

# 打开1.xlsx文件
workbook = load_workbook('SC.xlsx')

# 获取工作簿中所有工作表的名称
sheet_names = workbook.sheetnames

# 遍历每个工作表
for sheet_name in sheet_names:
    # 获取当前工作表
    sheet = workbook[sheet_name]
    
    # 读取当前工作表中的数据
    data = sheet.values
    rows = list(data)
    
    # 遍历每一行，检查是否所有单元格都不包含K1字符，如果是则删除该行
    filtered_rows = [row for row in rows if any('K1' in str(cell) for cell in row)]
    
    # 清空当前工作表
    sheet.delete_rows(1, sheet.max_row)
    
    # 将筛选后的行写入当前工作表
    for row_index, row_data in enumerate(filtered_rows, start=1):
        for column_index, cell_data in enumerate(row_data, start=1):
            sheet.cell(row=row_index, column=column_index, value=cell_data)

# 保存修改后的Excel文件
workbook.save('SC_modified.xlsx')

