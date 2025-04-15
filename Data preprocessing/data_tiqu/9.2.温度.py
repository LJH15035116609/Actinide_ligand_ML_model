import openpyxl
import re

def extract_contents_from_D(sheet):
    extracted_contents = []
    for cell in sheet['D']:
        cell_value = cell.value
        if isinstance(cell_value, str):
            matches = re.findall(r'(-?\d+(\.\d+)?)', cell_value)  # 使用正则表达式匹配数字
            for match in matches:
                content = match[0].strip()  # 获取匹配到的数字
                if content and '℃' not in cell_value:  # 确保内容非空且不包含℃符号
                    try:
                        content = float(content)  # 尝试将内容转换为数字格式
                        extracted_contents.append(content)
                    except ValueError:
                        pass

    return extracted_contents

# 打开 Excel 文件
workbook = openpyxl.load_workbook('SC_modified3 - 副本.xlsx')

# 创建一个新的 Excel 文件来保存提取的数据
output_workbook = openpyxl.Workbook()

# 遍历每个 sheet
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    extracted_contents = extract_contents_from_D(sheet)

    # 创建一个新的工作表来保存提取的数据
    output_sheet = output_workbook.create_sheet(title=sheet_name)

    # 将提取的数据写入到新的工作表中
    for content in extracted_contents:
        output_sheet.append([content])

# 删除默认的工作表
output_workbook.remove(output_workbook.active)

# 保存新的 Excel 文件
output_workbook.save('℃.xlsx')
